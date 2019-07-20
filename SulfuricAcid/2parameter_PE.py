# -*- coding: utf-8 -*-
import numpy as np
import matplotlib.pyplot as plt
from scipy.integrate import odeint
import math
from scipy.optimize import minimize
from openpyxl import Workbook, load_workbook
import pandas as pd
import scipy.stats as stats
import os
import openpyxl

def onlinePE(filename):# you need to give the file name in which experimental results are saved
    os.chdir('C:/Users/User/Documents/LabVIEW Data/2015 Projects/Combined')
    wb=load_workbook(filename)
    ws=wb.active 
    
    Temp=[]
    InletC=[]
    Flow=[]
    BAC=[]
    EBC=[]
    parameter1 = []
    parameter2 = []
    chisq = []
    ref_chisq = []
    conf_int1 = []
    conf_int2 = []
    t_value1 = []
    t_value2 = []
    t_value_ref = []
    
    #Reading Files
    row_count = ws.max_row
    for i in range(2,row_count):
        Temp.append(ws['A%d'%i].value)
        InletC.append(ws['C%d'%i].value)
        Flow.append(ws['B%d'%i].value)      # performed/performing experiments & measurements
        BAC.append(ws['D%d'%i].value)
        EBC.append(ws['E%d'%i].value)
        
        
    sigma_BAC = 0.03 # estimate of variance
    sigma_EBC = 0.0165
    sigma = [sigma_BAC, sigma_EBC]
    R = 8.314 # gas constant
    n_theta = 2                            # number of unknown model parameters
    n_y = 2                                # number of measured responses
    n_phi = 3                              # number of design variables
    theta0 = np.zeros(n_theta) 
    theta0 = [100, 90000] 
    theta = np.zeros(len(theta0))
    ntheta = np.ones(len(theta0)) # normaised value of parameter, this is what we actually estimate

            # controls
    
    u0 = np.zeros(n_phi)
    u0 = [100, 10, 1]
    u = np.zeros(len(u0))
    un = np.ones(len(u0))
    u1 = 140 # Temperature (K)
    u2 = 20 # Flowrate (microL/min)
    u3 = 1.5 # Initial concentration (mol/L)
    u = [u1, u2, u3] #controls or inputs
    inlet = 0
    bed_length = 200 # cm
    Ac = 4.91 * (10 ** -4) # cm2 # area of tube
    exp_data = pd.read_excel(filename)
    performed_exp = exp_data[['Temperature', 'Flowrate', 'Concentration']].values   # experimental conditions
    y_m = exp_data[['BAC', 'EBC']].values   # experiment measurements
    number_read=len(y_m)

    def normalised_conditions(performed_exp):
        norm_data = np.zeros([len(performed_exp),len(performed_exp[0,:])])
        for i in range(len(performed_exp)):
            norm_data[i,0] = performed_exp[i,0]/u0[0]
            norm_data[i,1] = performed_exp[i,1]/u0[1]
            norm_data[i,2] = performed_exp[i,2]/u0[2]
        return norm_data
    normalised_exp_conditions = normalised_conditions(performed_exp)
    #return normalised_exp_conditions
    
    def kinetic_model(x,z,un,ntheta):
        CA = x[0]
        CB = x[1]
        for i in range(len(theta0)):
            theta[i] = theta0[i] * ntheta[i]
        k1 = theta[0] * math.exp(-theta[1]/(R * ((u0[0]*un[0])+273.15)))   
        dAdz = (-k1 * (CA)) * (Ac/((u0[1]*un[1])/1000))
        dBdz = k1 * (CA) * (Ac/((u0[1]*un[1])/1000))       
        return [dAdz, dBdz]
    
        
    def loglikelihood(ntheta,un):
        solutions = np.zeros([2,2])
        for un in normalised_exp_conditions:
            soln = odeint(kinetic_model,[(u0[2]*un[2]),0],[0,200], mxstep = 3000, args=(un,ntheta))
            CA = soln[:, 0]
            CB = soln[:, 1]
            C = [CA, CB]
            solutions= np.append(solutions,C,axis=0)
        responses_new = solutions[2:,]
        yA = responses_new[0::2]
        yB = responses_new[1::2]
        yA_p = yA[:,-1] 
        yB_p = yB[:,-1]
        rho_1 = y_m[:,0] - yA_p 
        rho_2 = y_m[:,1] - yB_p
        rho = (rho_1/sigma[0])**2+(rho_2/sigma[1])**2
        residuals = np.sum(rho)
        neg_loglikelihood = math.log(2*math.pi) + 0.5 * (math.log(sigma[0]**2) + math.log(sigma[1]**2)) + 0.5 * residuals
        obj_fun = 0.5 * residuals
        return obj_fun
        #return yA_p,yB_p,residuals
    #return loglikelihood(ntheta,un)
    
    def parameter_estimation(ntheta,un):
        new_estimate = minimize(loglikelihood,[0.5,0.7],method = 'Nelder-Mead', options = {'maxiter':2000}, args=(un,))
        return new_estimate
    estimate = parameter_estimation(ntheta,un).x
    param1 = estimate[0]*theta0[0]
    param2 = estimate[1]*theta0[1]
    #return param1,param2

        ## Adequacy test/chisquare #

    alpha = 0.05
    conf_level = 1 - alpha
    dof = (((len(performed_exp)) * n_y) - len(theta0))
    def chisquare_test(performed_exp,conf_level):
        ref_chisquare = stats.chi2.ppf((conf_level),dof)
        chisquare_value = parameter_estimation(ntheta,un).fun
        p_value = 1 - stats.chi2.cdf(chisquare_value, dof)
        return ref_chisquare, chisquare_value
    wt_residuals = 2 * (chisquare_test(performed_exp,conf_level)[1])
    chisq_ref = chisquare_test(performed_exp,conf_level)[0]
    #return estimate, wt_residuals, chisq_ref

    ## Actual MBDoE starts from here ##

    epsilon = 0.01
    def perturbation(epsilon,estimate):      
        perturbated_matrix = np.zeros([len(estimate)+1,len(estimate)])
        for j in range(len(estimate)):
            for k in range(len(estimate)):
                if j==k:
                    perturbated_matrix[j,k] = estimate[j] * (1 + epsilon)
                else:
                    perturbated_matrix[j,k] = estimate[k]
        for j in range(len(estimate)):
            perturbated_matrix[-1,j] = estimate[j]
        return perturbated_matrix
    
    #return perturbation(epsilon, estimate)

    def sensitivity(un,ntheta,epsilon):
        solutions_sen = np.zeros([2,2])
        for ntheta in perturbation(epsilon,estimate):
            soln = odeint(kinetic_model,[(u0[2]*un[2]),0], [0,200], args=(un,ntheta))
            CA = soln[:, 0]
            CB = soln[:, 1]
            C = [CA, CB]
            solutions_sen= np.append(solutions_sen,C,axis=0)
        responses_sen = solutions_sen[n_y:,]
        y1_sen = responses_sen[0::n_y]
        y2_sen = responses_sen[1::n_y]
        y_sen = np.zeros([len(ntheta)+1,len(responses_sen)])
        y_sen = np.transpose([y1_sen[:,1],y2_sen[:,1]])
        sensitivity_matrix = np.zeros([len(estimate),n_y])
        for j in range(len(estimate)):
            for k in range(n_y):
                sensitivity_matrix[j,k] = ((y_sen[j,k] - y_sen[-1,k])/(epsilon))
        return sensitivity_matrix
    
    def information(un,ntheta):
        Fisher=np.zeros([len(theta0),len(theta0)])
        for j in range(len(sensitivity(un,ntheta,epsilon)[0,:])):
            Fisher = Fisher + (1/(sigma[j]**2)) * np.outer(sensitivity(un,ntheta,epsilon)[:,j],sensitivity(un,ntheta,epsilon)[:,j])
        return Fisher
    
    def obs_Fisher(performed_exp,sigma,estimate):
        obs_information = np.zeros([len(performed_exp),len(theta0),len(theta0)])
        for j in range(len(performed_exp)):
            obs_information[j,:,:] = information(normalised_exp_conditions[j,:],estimate)
        overall_obs_Fisher = np.zeros([len(theta0),len(theta0)])
        for j in range(len(performed_exp)):
            overall_obs_Fisher = overall_obs_Fisher + obs_information[j,:,:]
        return overall_obs_Fisher
    
    def obs_covariance(un,ntheta):
        obs_variance_matrix = np.linalg.inv(obs_Fisher(performed_exp,sigma,estimate))
        return obs_variance_matrix
    
    #return obs_covariance(un,ntheta)

    def correlation(ntheta):
        correlationmatrix = np.zeros([len(obs_covariance(un,ntheta)),len(obs_covariance(un,ntheta))])
        for i in range(len(obs_covariance(un,ntheta))):
            for j in range(len(obs_covariance(un,ntheta))):
                correlationmatrix[i,j] = obs_covariance(un,ntheta)[i,j]/(np.sqrt(obs_covariance(un,ntheta)[i,i] * obs_covariance(un,ntheta)[j,j]))
        return correlationmatrix

    #return correlation(ntheta)

    def t_test(ntheta, conf_level, dof):  
        variances = np.zeros(len(theta0))
        t_values = np.zeros(len(theta0))
        conf_interval = np.zeros(len(theta0))
        for j in range(len(theta0)):
            conf_interval[j] = np.sqrt(obs_covariance(un,ntheta)[j,j]) * estimate[j] * stats.t.ppf((1 - (alpha/2)), dof)
            t_values[j] = estimate[j]/(conf_interval[j])
        t_ref = stats.t.ppf((1-alpha),dof)
        return conf_interval,t_values,t_ref
    t_val1 = t_test(ntheta,conf_level,dof)[1][0]
    t_val2 = t_test(ntheta,conf_level,dof)[1][1]
    t_ref_val = t_test(ntheta,conf_level,dof)[2]
    c_int1 = t_test(ntheta,conf_level,dof)[0][0]
    c_int2 = t_test(ntheta,conf_level,dof)[0][1]
    
    #loading files
    wbwrite=openpyxl.load_workbook(filename)
    #print (wb.sheetnames)
    wswrite=wbwrite.active #opens the first sheet in the wb

    row_count = wswrite.max_row
    wswrite['F'+str(row_count)]=param1 #using the excel numbering system    
    wswrite['G'+str(row_count)]=param2 #using the excel numbering system
    wswrite['H'+str(row_count)]=wt_residuals #using the excel numbering system    
    wswrite['I'+str(row_count)]=chisq_ref #using the excel numbering system
    wswrite['J'+str(row_count)]=c_int1 #using the excel numbering system
    wswrite['K'+str(row_count)]=c_int2 #using the excel numbering system
    wswrite['L'+str(row_count)]=t_val1 #using the excel numbering system
    wswrite['M'+str(row_count)]=t_val2 #using the excel numbering system
    wswrite['N'+str(row_count)]=t_ref_val #using the excel numbering system
    
    #Saving File
    wbwrite.save(filename)# overwrites without warning. So be careful
    
    
    return (number_read, param1, param2, wt_residuals, chisq_ref, c_int1, c_int2, t_val1, t_val2, t_ref_val)
    #def exp_sen(un,ntheta):
        #sen1 = sensitivity(un,estimate,epsilon)
        #return sen1
    
    #def exp_Fisher(un,ntheta):
        #Fisher1 = information(un,estimate)
        #return Fisher1
    
    #def exp_covariance(un,ntheta):
        #variance_covariance_matrix = np.linalg.inv(obs_Fisher(performed_exp,sigma,estimate) + exp_Fisher(un,ntheta))
        #return variance_covariance_matrix
    
    #def D_optimal(un,ntheta):
        #obj_D_optimal = np.linalg.det(exp_covariance(un,ntheta))
        #return obj_D_optimal

    #def experiment_design(un,ntheta):
        #new_design = minimize(D_optimal,[1,1,1],method = 'SLSQP', bounds = ([0.7,1.5],[0.7,4],[0.1,2]), options = {'maxiter':10000, 'ftol':1e-20}, args = (ntheta,))
        #return new_design 
    #new_design = experiment_design(un,ntheta).x
    #new_Temp = new_design[0]
    #new_Flow = new_design[1]
    #new_Concentration = new_design[2]
    #return new_design     
    
        