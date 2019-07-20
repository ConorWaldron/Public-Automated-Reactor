# -*- coding: utf-8 -*-
"""
Created on Mon Jan 15 14:34:09 2018

@author: Conor
"""

import openpyxl
import os

def AppendResultFile(Filename, Temp, Flowrate, FeedC):#you need to give the file name in '', but LabVIEW does this automatically
    os.chdir('C:/Users/User/Documents/LabVIEW Data/2015 Projects/Combined')
    
    #loading files
    wb=openpyxl.load_workbook(Filename)
    #print (wb.sheetnames)
    ws=wb.active #opens the first sheet in the wb
    
    #Mainipulating Files
    row_count = ws.max_row
    ws.append([Temp, Flowrate, FeedC])
    #ws['A2']=Temp #using the excel numbering system    
    #ws['B2']=Flowrate #using the excel numbering system    
    #ws['C2']=FeedC #using the excel numbering system    
    
    #Saving File
    wb.save(Filename)# overwrites without warning. So be careful
    return 