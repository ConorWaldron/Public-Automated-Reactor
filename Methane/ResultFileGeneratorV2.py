# -*- coding: utf-8 -*-
"""
Created on Fri May 31 2019

@author: Conor
"""

import openpyxl
import os

def AppendResultFile(FilenameByte, Temp, FlowO2, FlowN2, FlowCh4He, Catmass):#you need to give the file name in '', but LabVIEW does this automatically
    os.chdir('C:/Users/Solomon/Documents/Experimental conditions')
    
    Filename=FilenameByte.decode('ASCII')
    
    #loading files
    wb=openpyxl.load_workbook(Filename)
    #print (wb.sheetnames)
    ws=wb.active #opens the first sheet in the wb
    
    #Mainipulating Files
    #row_count = ws.max_row
    ws.append([Temp, FlowO2, FlowN2, FlowCh4He, Catmass])
    #ws['A2']=Temp #using the excel numbering system    
    #ws['B2']=Flowrate #using the excel numbering system    
    #ws['C2']=FeedC #using the excel numbering system    
    
    #Saving File
    wb.save(Filename)# overwrites without warning. So be careful
    return 