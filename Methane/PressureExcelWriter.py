# -*- coding: utf-8 -*-
"""
Created on Fri May 31 2019

@author: Conor
"""

import openpyxl
import os

def AppendPressures(FilenameByte, Pin, Pout):#you need to give the file name in '', but LabVIEW does this automatically
    os.chdir('C:/Users/Solomon/Documents/Experimental conditions')
    
    Filename=FilenameByte.decode('ASCII')
    
    #loading files
    wb=openpyxl.load_workbook(Filename)
    ws=wb.active #opens the first sheet in the wb
    
    #Mainipulating Files
    row_count = ws.max_row
    ws['F'+str(row_count)]=Pin #using the excel numbering system    
    ws['G'+str(row_count)]=Pout #using the excel numbering system
        
    #Saving File
    wb.save(Filename)# overwrites without warning. So be careful   
        
    return 