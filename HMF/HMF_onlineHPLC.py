#Code to extract the most recent HPLC measurment for the HPLC system

import os
import openpyxl 
import csv, numpy

def HPLC(HPLC_Createdfile, MyRecordFile):
    CalibrationFactors=[50463, 72301, 247669, 59340] #"HMF", "HFCA", "FCA", "FDCA"
    SpeciesAreas = [0, 0, 0, 0, 0] #"HMF", "HFCA", "FCA", "FDCA", "Unknown"
    SpeciesConc = [0, 0, 0, 0, 0]
        
    os.chdir('C:/Users/User/Documents/LabVIEW Data/2015 Projects/Combined')

#This code only works with new excel format .xlsx, but HPLC can only make either .csv or .xls
    #wb=openpyxl.load_workbook(HPLC_Createdfile)
    #ws=wb.active  
    #Reading Files
    #row_count = ws.max_row
    #for i in range(2,row_count+1):
    #    AqDate.append(ws['D%d'%i].value)
    #    PeakName.append(ws['E%d'%i].value)
    #    Area.append(ws['K%d'%i].value)   
    
#CSV data loading method
    CWDfiles=os.listdir("./")#returns list of files in CWD
    #print ("All items in the CWD incldue "+str(CWDfiles))#this is how you combine text with variable strings in Python
    for entry in CWDfiles:
        root, ext = os.path.splitext(entry)
        #print(root) #shows just the root (name) of each file in CWD
        #print(ext)
        if ext == ".csv" and HPLC_Createdfile in entry:#here sample is the target file name, it needs to be changed for new files.
            AreaRead = []
            AqDateRead = []
            PeakNameRead = []
            with open(entry, 'r') as inputfile:
                reader = csv.reader(inputfile)
                full_file = list(reader)
                #print reader#not sure what type of variable reader is
                #print full_file #full file is a matrix with all the entries from the excell file, so it is the full file...
            
    #we start at 1, as row 0 is the header
    for i in range (1, len(full_file)):
        AreaRead.append (full_file[i][10])
        AqDateRead.append (full_file[i][3])
        PeakNameRead.append (full_file[i][4])
    inputfile.close()

    AreaNP=numpy.array(AreaRead,dtype=float)
    AqDate=numpy.array(AqDateRead)
    PeakName=numpy.array(PeakNameRead)


    #Need to identify the last sample
    LastTime=AqDate[-1]
    LastSampleIndicies = []
    for i in range(0,len(AreaNP)):
        if AqDate[i]==AqDate[-1]:
            LastSampleIndicies.append(i)
    
    #Need to identify which ones are which species
    ListofSpecies=["HMF", "HFCA", "FCA", "FDCA", "Unknown"]
    for i in range(0, len(LastSampleIndicies)):
        for j in range(0, len(ListofSpecies)):
            if PeakName[LastSampleIndicies[i]]==ListofSpecies[j]:
                #print("true")
                SpeciesAreas[j]=AreaNP[LastSampleIndicies[i]]
            #else:
                #print("no match")
                
    for i in range(0,4):
        SpeciesConc[i] = SpeciesAreas[i]/CalibrationFactors[i]
        
    #append results to a different file
    #loading files
    wbwrite=openpyxl.load_workbook(MyRecordFile)
    #print (wb.sheetnames)
    wswrite=wbwrite.active #opens the first sheet in the wb
    
    #Mainipulating Files
    row_count = wswrite.max_row
    wswrite['I'+str(row_count)]=SpeciesConc[0] #using the excel numbering system    
    wswrite['J'+str(row_count)]=SpeciesConc[1] #using the excel numbering system 
    wswrite['K'+str(row_count)]=SpeciesConc[2] #using the excel numbering system    
    wswrite['L'+str(row_count)]=SpeciesConc[3] #using the excel numbering system
    #wswrite['M'+str(row_count)]=SpeciesAreas[4] #using the excel numbering system

    #Saving File
    wbwrite.save(MyRecordFile)# overwrites without warning. So be careful
    
    return SpeciesConc[0], SpeciesConc[1], SpeciesConc[2], SpeciesConc[3] 